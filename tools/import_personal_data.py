from __future__ import annotations

import csv
import difflib
import json
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
ARCHIVO_PERSONAL = ROOT / "archivo-personal"
IMPORTACIONES_DIR = ARCHIVO_PERSONAL / "importaciones"
EXTRACTED_WRITINGS_DIR = IMPORTACIONES_DIR / "escritos-extraidos"
FULL_OUTPUT_PATH = IMPORTACIONES_DIR / "private-life-import-full.json"
CURATED_OUTPUT_PATH = IMPORTACIONES_DIR / "private-life-import-curado.json"
HABITS_DIR = ARCHIVO_PERSONAL / "habitos" / "Loop Habits CSV 2026-04-09"
SERIES_CSV = ARCHIVO_PERSONAL / "series" / "46906f9a-6a9f-4101-bc43-67a5ba7cd38d.csv"
WRITINGS_DIR = ARCHIVO_PERSONAL / "escritos"
LIFE_XLSX = ARCHIVO_PERSONAL / "peliculas" / "Life .xlsx"
SERIES_TYPES = {"TV Series", "TV Mini Series"}
MOVIE_TYPES = {
    "Movie",
    "movie",
    "Feature Film",
    "feature",
    "TV Movie",
    "tvMovie",
    "TV Special",
    "Video",
    "Short",
    "short",
}
MOVIE_TITLE_EQUIVALENTS = {
    "threesome": ["the threesome"],
    "happiness for beginner": ["happiness for beginners"],
    "glass onion a knives out mystery": ["glass onion"],
    "love at second sight mon inconnue": ["love at second sight", "mon inconnue"],
    "koto no ha no niwa": ["kotonoha no niwa", "the garden of words"],
    "quattro metà": ["four to dinner", "4 metà"],
    "the hanting game": ["the hating game"],
    "to all the boys ive loved befor": ["to all the boys ive loved before"],
    "violet y finch": ["all the bright places"],
    "clara y claire": ["who you think i am", "celle que vous croyez"],
    "asesinato en el orient express": ["murder on the orient express"],
    "misanthrope": ["to catch a killer"],
    "heojil kyolshimdecision to leave": ["decision to leave", "heojil kyolshim"],
    "amor de película": ["the big love picture"],
    "godzilla king of the monsters!": ["godzilla king of the monsters"],
    "men in black 3": ["men in black3"],
    "furious seven": ["furious 7", "fast and furious 7"],
    "21 blackjack": ["21"],
    "wolverine": ["the wolverine"],
    "x men 2": ["x2", "x2 x men united"],
    "men in black 2": ["men in black ii"],
    "men in black 3": ["men in black iii", "men in black 3"],
    "the boy in the striped pyjamas": ["the boy in the striped pajamas"],
    "spy kids all the time in the world in 4d": ["spy kids 4 all the time in the world", "spy kids all the time in the world"],
}

TEXTOS_TITLES = [
    "Tiempo",
    "El Diablo",
    "Mejorar Excusandose",
    "El Mundo es un Pendulo",
]

THOUGHT_PERIOD_START = datetime(2024, 9, 3)

DATE_HEADER_RE = re.compile(r"(?m)^(?:(\d{1,2}:\d{2})\s+)?(\d{1,2}/\d{1,2}/\d{2,4})\b")
COSITAS_HEADER_RE = re.compile(
    r"(?mi)^(el\s+(?:domingo|lunes|martes|miercoles|miércoles|jueves|viernes|sabado|sábado))\s*:"
)


def clean_text(value: object | None) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value or "").strip()
    replacements = {
        "³": "3",
        "ł": "3",
        "BaÃ±o": "Baño",
        "BaÃƒÂ±o": "Baño",
        "FÃ­sico": "Físico",
        "FÃƒÂ­sico": "Físico",
        "pelÃ­cula": "película",
        "pelÃƒÂ­cula": "película",
        "miÃ©rcoles": "miércoles",
        "sÃ¡bado": "sábado",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return " ".join(text.split())


def slugify_tag(value: str) -> str:
    base = clean_text(value).lower()
    for source, target in {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
        "?": "",
        ",": "",
        ".": "",
        ":": "",
        "/": "-",
        " ": "-",
    }.items():
        base = base.replace(source, target)
    return base.strip("-")


def sort_entries(entries: Iterable[dict]) -> list[dict]:
    return sorted(entries, key=lambda entry: (entry["date"], entry["id"]), reverse=True)


def normalize_date(raw: object | None) -> tuple[str | None, bool]:
    value = clean_text(raw)
    if not value:
      return None, False

    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", value):
        day, month, year = value.split("/")
        return f"{year}-{month}-{day}", False
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2}", value):
        day, month, year = value.split("/")
        return f"20{int(year):02d}-{int(month):02d}-{int(day):02d}", False
    if re.fullmatch(r"\d{1,2}/\d{4}", value):
        month, year = value.split("/")
        return f"{year}-{int(month):02d}-01", True
    if re.fullmatch(r"-?\d{4}", value):
        return f"{value[-4:]}-01-01", True
    return None, False


def excel_date_to_iso(value: object) -> tuple[str | None, bool]:
    if value is None:
        return None, False
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d"), False
    if isinstance(value, (int, float)):
        year = int(value)
        if 1000 <= year <= 2100:
            return f"{year}-01-01", True
    return normalize_date(value)


def is_childhood_marker(value: object) -> bool:
    return isinstance(value, (int, float)) and float(value) < 0


def normalize_habit_name(name: str) -> str:
    normalized = clean_text(name).rstrip("?").strip()
    lowered = normalized.lower()
    if lowered in {"bao", "bano", "baño"}:
        return "Baño"
    if lowered in {"ejercico fsico", "ejercicio fisico", "ejercicio físico"}:
        return "Ejercicio físico"
    return normalized


def parse_habits_full() -> list[dict]:
    entries: list[dict] = []
    if not HABITS_DIR.exists():
        return entries

    for habit_dir in sorted(HABITS_DIR.iterdir()):
        if not habit_dir.is_dir():
            continue

        base_name = habit_dir.name.split(" ", 1)[1] if " " in habit_dir.name else habit_dir.name
        habit_name = normalize_habit_name(base_name)
        checkmarks_path = habit_dir / "Checkmarks.csv"
        if not checkmarks_path.exists():
            continue

        with checkmarks_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if not (row.get("Value") or "").strip().startswith("YES"):
                    continue

                date = (row.get("Date") or "").strip()
                if not date:
                    continue

                notes = clean_text(row.get("Notes"))
                content = f"Registro diario de {habit_name}."
                if notes:
                    content = f"{content} Nota: {notes}"

                entries.append(
                    {
                        "id": f"habit-{slugify_tag(habit_name)}-{date}",
                        "type": "habit",
                        "section": "habit",
                        "title": habit_name,
                        "content": content,
                        "date": date,
                        "tags": ["habit", "loop", slugify_tag(habit_name)],
                    }
                )

    return sort_entries(entries)


def split_genres(value: object | None) -> list[str]:
    text = clean_text(value)
    if not text:
        return []

    normalized = (
        text.replace(" y ", ",")
        .replace(" Y ", ",")
        .replace(" e ", ",")
        .replace(" E ", ",")
    )

    genres: list[str] = []
    seen: set[str] = set()
    for item in normalized.split(","):
        tag = slugify_tag(item)
        if not tag or tag in seen:
            continue
        genres.append(tag)
        seen.add(tag)
    return genres


def find_imdb_exports() -> list[Path]:
    candidates: list[Path] = []
    for path in ARCHIVO_PERSONAL.rglob("*.csv"):
        lowered = str(path).lower()
        if "\\habitos\\" in lowered or "\\importaciones\\" in lowered:
            continue
        candidates.append(path)
    return sorted(candidates)


def parse_imdb_rows(allowed_types: set[str]) -> list[dict]:
    entries: list[dict] = []

    for csv_path in find_imdb_exports():
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            fields = set(reader.fieldnames or [])
            if "Title" not in fields or "Title Type" not in fields:
                continue

            for row in reader:
                title_type = clean_text(row.get("Title Type"))
                if title_type not in allowed_types:
                    continue

                title = clean_text(row.get("Title"))
                date = (row.get("Date Rated") or row.get("Modified") or row.get("Created") or "").strip()
                if not title or not date:
                    continue

                entries.append(
                    {
                        "title": title,
                        "original_title": clean_text(row.get("Original Title")),
                        "date": date,
                        "your_rating": clean_text(row.get("Your Rating")),
                        "imdb_rating": clean_text(row.get("IMDb Rating")),
                        "year": clean_text(row.get("Year")),
                        "genres": split_genres(row.get("Genres")),
                        "runtime": clean_text(row.get("Runtime (mins)")),
                    }
                )

    return entries


def parse_series() -> list[dict]:
    entries: list[dict] = []

    for row in parse_imdb_rows(SERIES_TYPES):
        details = [
            f"IMDb {row['imdb_rating']}." if row["imdb_rating"] else "",
            f"{row['year']}." if row["year"] else "",
            f"{row['runtime']} min." if row["runtime"] else "",
        ]
        tags = ["series", "imported", *row["genres"][:4]]
        if row["your_rating"]:
            tags.append("rated")

        entries.append(
            {
                "id": f"series-{slugify_tag(row['title'])}-{row['date']}",
                "type": "series",
                "section": "series",
                "title": row["title"],
                "content": " ".join(part for part in details if part),
                "date": row["date"],
                "rating": row["your_rating"] or None,
                "tags": tags,
            }
        )

    return sort_entries(entries)


def parse_life_movies_xlsx() -> list[dict]:
    entries: list[dict] = []
    if not LIFE_XLSX.exists():
        return entries

    wb = load_workbook(LIFE_XLSX, data_only=True)
    if "Peliculas" not in wb.sheetnames:
        return entries

    ws = wb["Peliculas"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = clean_text(row[0] if len(row) > 0 else None)
        genre_year = clean_text(row[4] if len(row) > 4 else None)
        raw_watched_at = row[7] if len(row) > 7 else None
        watched_at, approximate_date = excel_date_to_iso(raw_watched_at)
        childhood_date = is_childhood_marker(raw_watched_at)
        reaction = clean_text(row[8] if len(row) > 8 else None).lower()

        if not title or not genre_year or not watched_at:
            continue

        genre_match = re.match(r"(.+?)\s+(-?\d{4})$", genre_year)
        if genre_match:
            genres_text = clean_text(genre_match.group(1))
            release_year = genre_match.group(2).replace("-", "")
        else:
            year_inline = re.search(r"(\d{4})", str(row[7] if len(row) > 7 else ""))
            if not year_inline:
                continue
            genres_text = genre_year
            release_year = year_inline.group(1)

        reaction_tag = ""
        reaction_label = ""
        if "wow" in reaction:
            reaction_tag = "wow"
            reaction_label = "wow"
        elif "like" in reaction:
            reaction_tag = "liked"
            reaction_label = "i like"

        tags = ["movie", "imported", "life-xlsx", *split_genres(genres_text)]
        if reaction_tag:
            tags.append(reaction_tag)
        if approximate_date:
            tags.append("approx-date")
        if childhood_date:
            tags.append("childhood")

        content_parts = [f"{genres_text}.", f"{release_year}.", f"{reaction_label}." if reaction_label else ""]
        entries.append(
            {
                "id": f"movie-{slugify_tag(title)}-{watched_at}",
                "type": "movie",
                "section": "movie",
                "title": title,
                "content": " ".join(part for part in content_parts if part),
                "date": watched_at,
                "release_year": release_year,
                "tags": tags[:8],
            }
        )

    return sort_entries(entries)


def parse_imdb_movies() -> list[dict]:
    entries: list[dict] = []

    for row in parse_imdb_rows(MOVIE_TYPES):
        details = [
            f"IMDb {row['imdb_rating']}." if row["imdb_rating"] else "",
            f"{row['year']}." if row["year"] else "",
            f"{row['runtime']} min." if row["runtime"] else "",
        ]
        tags = ["movie", "imported", "imdb", *row["genres"][:4]]
        if row["your_rating"]:
            tags.append("rated")

        entries.append(
            {
                "id": f"movie-imdb-{slugify_tag(row['title'])}-{row['date']}",
                "type": "movie",
                "section": "movie",
                "title": row["title"],
                "original_title": row.get("original_title"),
                "content": " ".join(part for part in details if part),
                "date": row["date"],
                "release_year": row["year"],
                "rating": row["your_rating"] or None,
                "tags": tags,
            }
        )

    return sort_entries(entries)


def movie_match_key(title: str) -> str:
    base = clean_text(title).lower()
    replacements = {
        "&": "and",
        ":": "",
        "'": "",
        ".": "",
        ",": "",
        "(": "",
        ")": "",
        "-": " ",
    }
    for source, target in replacements.items():
        base = base.replace(source, target)
    return " ".join(base.split())


def movie_match_aliases(title: str) -> set[str]:
    base = movie_match_key(title)
    aliases = {base}
    variants = [
        re.sub(r"^007\s+", "", base),
        re.sub(r"\(.*?\)", "", base),
        re.sub(r"\bthe movie\b", "", base),
        re.sub(r"\bpart two\b", "part 2", base),
        re.sub(r"\bpart 2\b", "part two", base),
        re.sub(r"\biii\b", "3", base),
        re.sub(r"\b3\b", "iii", base),
    ]
    for variant in variants:
        cleaned = " ".join(variant.split())
        if cleaned:
            aliases.add(cleaned)
    if ":" in base:
        aliases.add(base.split(":")[0].strip())
    if base.startswith("the "):
        aliases.add(base.removeprefix("the ").strip())
    if base.startswith("007 "):
        aliases.add(base.removeprefix("007 ").strip())
    for extra in MOVIE_TITLE_EQUIVALENTS.get(base, []):
        aliases.add(movie_match_key(extra))
    return {alias for alias in aliases if alias}


def merge_movie_sources(life_entries: list[dict], imdb_entries: list[dict]) -> list[dict]:
    if not imdb_entries:
        return life_entries

    imdb_by_key: dict[str, dict] = {}
    for entry in imdb_entries:
        alias_titles = [entry["title"]]
        if entry.get("original_title"):
            alias_titles.append(entry["original_title"])
        keys = {alias for alias_title in alias_titles for alias in movie_match_aliases(alias_title)}
        for key in keys:
            imdb_by_key.setdefault(key, entry)

    merged: list[dict] = []
    used_keys: set[str] = set()
    unmatched_imdb = list(imdb_entries)

    for life_entry in life_entries:
        life_keys = movie_match_aliases(life_entry["title"])
        imdb_entry = next((imdb_by_key[key] for key in life_keys if key in imdb_by_key), None)
        if not imdb_entry:
            best_match = None
            best_score = 0.0
            for candidate in unmatched_imdb:
                if life_entry.get("release_year") and candidate.get("release_year"):
                    if str(life_entry["release_year"]) != str(candidate["release_year"]):
                        continue
                score = max(
                    difflib.SequenceMatcher(None, life_key, imdb_key).ratio()
                    for life_key in life_keys
                    for imdb_key in {
                        alias
                        for alias_title in [candidate["title"], candidate.get("original_title") or ""]
                        if alias_title
                        for alias in movie_match_aliases(alias_title)
                    }
                )
                if score > best_score:
                    best_score = score
                    best_match = candidate

            if best_match and best_score >= 0.78:
                imdb_entry = best_match
            else:
                merged.append(life_entry)
                continue

        used_keys.add(imdb_entry["id"])
        unmatched_imdb = [entry for entry in unmatched_imdb if entry["id"] != imdb_entry["id"]]
        merged.append(
            {
                **life_entry,
                "title": imdb_entry["title"],
                "rating": imdb_entry.get("rating"),
                "content": " ".join(
                    part
                    for part in [
                        life_entry["content"],
                        imdb_entry["content"],
                    ]
                    if part
                ).strip(),
                "tags": list(dict.fromkeys([*life_entry["tags"], *imdb_entry["tags"]])),
            }
        )

    for imdb_entry in imdb_entries:
        if imdb_entry["id"] in used_keys:
            continue
        merged.append(imdb_entry)

    return sort_entries(merged)


def pick_preferred_movie_entry(current: dict, candidate: dict) -> dict:
    current_approx = "approx-date" in current.get("tags", [])
    candidate_approx = "approx-date" in candidate.get("tags", [])

    if current_approx != candidate_approx:
        return current if not current_approx else candidate

    current_rating = current.get("rating") not in (None, "")
    candidate_rating = candidate.get("rating") not in (None, "")
    if current_rating != candidate_rating:
        return current if current_rating else candidate

    return current if current["date"] >= candidate["date"] else candidate


def dedupe_movies(entries: list[dict]) -> list[dict]:
    deduped: dict[tuple[str, str], dict] = {}

    for entry in entries:
        release_year = str(entry.get("release_year") or "")
        key = (movie_match_key(entry["title"]), release_year)
        existing = deduped.get(key)
        if not existing:
            deduped[key] = entry
            continue

        preferred = pick_preferred_movie_entry(existing, entry)
        other = entry if preferred is existing else existing
        deduped[key] = {
            **preferred,
            "content": " ".join(
                part for part in [preferred.get("content", ""), other.get("content", "")] if part
            ).strip(),
            "tags": list(dict.fromkeys([*preferred.get("tags", []), *other.get("tags", [])])),
            "rating": preferred.get("rating") or other.get("rating"),
        }

    cleaned = []
    for entry in deduped.values():
        if entry["title"] == "Men in Black3":
            entry = {**entry, "title": "Men in Black 3"}
        cleaned.append(entry)

    return sort_entries(cleaned)


def parse_life_books_xlsx() -> list[dict]:
    entries: list[dict] = []
    if not LIFE_XLSX.exists():
        return entries

    wb = load_workbook(LIFE_XLSX, data_only=True)
    if "Libros" not in wb.sheetnames:
        return entries

    ws = wb["Libros"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = clean_text(row[0] if len(row) > 0 else None)
        author = clean_text(row[4] if len(row) > 4 else None)
        publish_year = clean_text(row[5] if len(row) > 5 else None)
        start_date, start_approx = excel_date_to_iso(row[7] if len(row) > 7 else None)
        end_date, end_approx = excel_date_to_iso(row[9] if len(row) > 9 else None)
        rating = row[12] if len(row) > 12 else None

        if not title:
            continue

        chosen_date = end_date or start_date
        approximate_date = end_approx if end_date else start_approx
        if not chosen_date:
            continue

        tags = ["book", "imported", "life-xlsx"]
        if rating not in (None, ""):
            tags.append("rated")
        if approximate_date:
            tags.append("approx-date")

        content_parts = [
            f"{author}." if author else "",
            f"{publish_year}." if publish_year else "",
            f"Inicio {start_date}." if start_date else "",
            f"Fin {end_date}." if end_date else "",
        ]
        entries.append(
            {
                "id": f"book-{slugify_tag(title)}-{chosen_date}",
                "type": "book",
                "section": "book",
                "title": title,
                "content": " ".join(part for part in content_parts if part),
                "date": chosen_date,
                "rating": rating if rating not in (None, "") else None,
                "tags": tags,
            }
        )

    return sort_entries(entries)


def parse_jw_milestones_xlsx() -> list[dict]:
    entries: list[dict] = []
    if not LIFE_XLSX.exists():
        return entries

    wb = load_workbook(LIFE_XLSX, data_only=True)
    if "Hitos JW" not in wb.sheetnames:
        return entries

    ws = wb["Hitos JW"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = clean_text(row[1] if len(row) > 1 else None)
        date_iso, approximate_date = excel_date_to_iso(row[3] if len(row) > 3 else None)
        note = clean_text(row[4] if len(row) > 4 else None)
        if not title or not date_iso:
            continue

        tags = ["jw", "milestone", "life-xlsx"]
        if approximate_date:
            tags.append("approx-date")

        entries.append(
            {
                "id": f"jw-{slugify_tag(title)}-{date_iso}",
                "type": "memory",
                "section": "general",
                "title": title,
                "content": f"{note}." if note else "",
                "date": date_iso,
                "tags": tags,
            }
        )

    return sort_entries(entries)


def normalize_extracted_text(text: str) -> str:
    lines = [clean_text(line) for line in text.splitlines()]
    return "\n".join(line for line in lines if line)


def collapse_content(text: str) -> str:
    parts = [clean_text(part) for part in text.splitlines()]
    return " ".join(part for part in parts if part)


def clip_content(text: str, limit: int = 1800) -> str:
    value = text.strip()
    if len(value) <= limit:
        return value
    return f"{value[:limit].rstrip()}..."


def file_date_string(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d")


def thought_fallback_date(index: int, offset: int = 0) -> str:
    return (THOUGHT_PERIOD_START + timedelta(days=index + offset)).strftime("%Y-%m-%d")


def normalize_date_header(raw: str) -> str | None:
    if not re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", raw):
        return None
    day, month, year = raw.split("/")
    if len(year) == 2:
        year = f"20{int(year):02d}"
    return f"{year}-{int(month):02d}-{int(day):02d}"


def split_solo_entries(text: str, path: Path) -> list[dict]:
    entries: list[dict] = []
    matches = list(DATE_HEADER_RE.finditer(text))
    if not matches:
        return entries

    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        block = text[start:end].strip()
        date_raw = match.group(2)
        date_iso = normalize_date_header(date_raw) or thought_fallback_date(index)
        entries.append(
            {
                "id": f"thought-solo-{date_iso}-{index}",
                "type": "note",
                "section": "thought",
                "title": f"Solo {date_raw}",
                "content": clip_content(collapse_content(block)),
                "date": date_iso,
                "tags": ["escritos", "thought", "solo"],
            }
        )

    return entries


def split_cositas_entries(text: str, path: Path) -> list[dict]:
    entries: list[dict] = []
    matches = list(COSITAS_HEADER_RE.finditer(text))
    if not matches:
        return entries

    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        block = text[start:end].strip()
        heading = clean_text(match.group(1)).title()
        inner_date_match = re.search(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b", block)
        date_iso = normalize_date_header(inner_date_match.group(1)) if inner_date_match else None
        approximate = date_iso is None
        date_iso = date_iso or thought_fallback_date(index, offset=12)
        tags = ["escritos", "thought", "cositas-xd"]
        if approximate:
            tags.append("approx-date")

        entries.append(
            {
                "id": f"thought-cositas-{date_iso}-{index}",
                "type": "note",
                "section": "thought",
                "title": f"{heading} {index + 1}" if approximate else heading,
                "content": clip_content(collapse_content(block)),
                "date": date_iso,
                "tags": tags,
            }
        )

    return entries


def split_textos_entries(text: str, path: Path) -> list[dict]:
    entries: list[dict] = []
    title_re = re.compile(rf"(?m)^({'|'.join(re.escape(title) for title in TEXTOS_TITLES)})\s*$")
    matches = list(title_re.finditer(text))
    if not matches:
        return entries

    fallback_date = file_date_string(path)
    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        title = clean_text(match.group(1))
        block = text[start:end].strip()
        entries.append(
            {
                "id": f"philosophy-{slugify_tag(title)}-{fallback_date}-{index}",
                "type": "note",
                "section": "philosophy",
                "title": title,
                "content": clip_content(collapse_content(block), limit=2400),
                "date": fallback_date,
                "tags": ["escritos", "philosophy", slugify_tag(title)],
            }
        )

    return entries


def parse_writings() -> list[dict]:
    entries: list[dict] = []
    if not WRITINGS_DIR.exists():
        return entries

    EXTRACTED_WRITINGS_DIR.mkdir(parents=True, exist_ok=True)

    for pdf_path in sorted(WRITINGS_DIR.glob("*.pdf")):
        if pdf_path.stem.lower().strip() == "life":
            continue

        reader = PdfReader(str(pdf_path))
        raw_text = "\n".join((page.extract_text() or "") for page in reader.pages)
        normalized = normalize_extracted_text(raw_text)
        if not normalized:
            continue

        (EXTRACTED_WRITINGS_DIR / f"{pdf_path.stem}.md").write_text(normalized, encoding="utf-8")

        stem = pdf_path.stem.lower().strip()
        if stem == "solo":
            entries.extend(split_solo_entries(normalized, pdf_path))
            continue
        if stem == "cositas xd":
            entries.extend(split_cositas_entries(normalized, pdf_path))
            continue
        if stem == "textos":
            entries.extend(split_textos_entries(normalized, pdf_path))
            continue

        entries.append(
            {
                "id": f"writing-{slugify_tag(pdf_path.stem)}-{file_date_string(pdf_path)}",
                "type": "note",
                "section": "thought",
                "title": clean_text(pdf_path.stem),
                "content": clip_content(collapse_content(normalized)),
                "date": file_date_string(pdf_path),
                "tags": ["escritos", slugify_tag(pdf_path.stem)],
            }
        )

    return sort_entries(entries)


def build_payload(entries: list[dict], note: str) -> dict:
    return {
        "version": 2,
        "exportedAt": None,
        "entries": sort_entries(entries),
        "meta": {
            "source": "archivo-personal",
            "notes": [note, "Archivo compatible con el importador JSON de la app."],
        },
    }


def main() -> None:
    habits_full = parse_habits_full()
    series_entries = parse_series()
    life_movie_entries = parse_life_movies_xlsx()
    imdb_movie_entries = parse_imdb_movies()
    movie_entries = dedupe_movies(merge_movie_sources(life_movie_entries, imdb_movie_entries))
    book_entries = parse_life_books_xlsx()
    jw_entries = parse_jw_milestones_xlsx()
    writing_entries = parse_writings()

    full_entries = [
        *habits_full,
        *series_entries,
        *movie_entries,
        *book_entries,
        *jw_entries,
        *writing_entries,
    ]

    IMPORTACIONES_DIR.mkdir(parents=True, exist_ok=True)
    FULL_OUTPUT_PATH.write_text(
        json.dumps(build_payload(full_entries, "Importación completa con histórico real."), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    CURATED_OUTPUT_PATH.write_text(
        json.dumps(build_payload(full_entries, "Importación curada para la app actual."), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    print(f"Generated {FULL_OUTPUT_PATH}")
    print(f"Generated {CURATED_OUTPUT_PATH}")
    print(
        "Counts:",
        json.dumps(
            {
                "habits_full": len(habits_full),
                "series": len(series_entries),
                "movies_life": len(life_movie_entries),
                "movies_imdb": len(imdb_movie_entries),
                "movies_merged": len(movie_entries),
                "books": len(book_entries),
                "jw_milestones": len(jw_entries),
                "writings": len(writing_entries),
            }
        ),
    )


if __name__ == "__main__":
    main()
